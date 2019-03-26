# -*- coding: utf-8 -*-
"""
    
    DF Funs

    for DB - ETL / API

Created on Tue Feb 12 18:29:58 2019
@author: Markus.Meister
"""
#%% -- imports --
import numpy as np
import pandas as pd
import dask
import dask.dataframe as dd
from openpyxl import load_workbook
from mpi4py import MPI
comm = MPI.COMM_WORLD
#%% -- main funs --

def write_df_to_excel(df, excel_name, sheet_name, startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(excel_name, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
    
    
    try:
        # try to open an existing workbook
        writer.book = load_workbook(excel_name)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    writer.close()

def dict_to_df(dict,thresholds=(0,np.inf),exceptions=[]):
    # finding the maximum no. elements
    max_elem = 0
    for d in dict:
        if len(dict[d]) > max_elem:
            max_elem = len(dict[d])
#    if max(thresholds) < max_elem:
#        max_elem = max(thresholds)
    # defining a cluster dict
    # loop over all groups
    # write all correponding KW into the group entry of the dict
    my_clusters = {}
    for c in dict:
        p = np.array(dict[c])
        psz = comm.allreduce(p.shape[0])
        if (psz < min(thresholds) or psz > max(thresholds)) and not c in exceptions:
            continue
        temp_cluster = (np.ones([max_elem]) * np.nan).astype(str)
        nan_slice = p.shape[0]
        temp_cluster[:nan_slice] = p
        my_clusters[c] = temp_cluster
    # return a dataframe from the dict
    return pd.DataFrame.from_dict(my_clusters)

def eval_value_dfs(df1,df2,vkeys,rkey):
    val_df = {a:np.zeros(vkeys.size,dtype=np.int32) for a in df1}
    for d in df1:
        for j,c in enumerate(vkeys):
            val_df[d][j] = df2.loc[df2[rkey].isin(df1[d]),c].values.sum()
    val_df['key'] = vkeys
    
    val_df = pd.DataFrame.from_dict(val_df).set_index('key')
    
    val_df_arr = comm.allreduce(val_df.values)
    
    val_df.loc[:,:] = val_df_arr
    
    return val_df

def mean_values_df(df):
    df2 = pd.DataFrame(columns=df.columns)
    df2.loc[0,:] = df.values.mean(axis=0)
    return df2

def dict_to_dd(dict,thresholds=(0,np.inf)):
    # finding the maximum no. elements
    max_elem = 0
    for d in dict:
        if len(dict[d]) > max_elem:
            max_elem = len(dict[d])
    if max(thresholds) < max_elem:
        max_elem = max(thresholds)
    # defining a cluster dict
    # loop over all groups
    # write all correponding KW into the group entry of the dict
    my_clusters = {}
    for c in dict:
        p = np.array(dict[c])
        psz = p.shape[0]
        if psz < min(thresholds) or psz > max(thresholds):
            continue
        temp_cluster = (np.ones([max_elem]) * np.nan).astype(str)
        nan_slice = p.shape[0]
        temp_cluster[:nan_slice] = p
        my_clusters[c] = temp_cluster
    # return a dataframe from the dict
    return dd.from_pandas(pd.DataFrame.from_dict(my_clusters),npartitions=4)

def eval_value_dds(df1,df2,vkeys,rkey):
    val_df = {a:np.zeros(vkeys.size,dtype=np.int32) for a in df1}
    for d in df1:
        for j,c in enumerate(vkeys):
            val_df[d][j] = df2.loc[df2[rkey].isin(df1[d]),c].values.sum().compute()
    val_df['key'] = vkeys
    
    val_df = dd.DataFrame.from_dict(val_df,npartitions=4).set_index('key')
    
    return val_df

def mean_values_dd(dd):
    dd2 = pd.DataFrame(columns=dd.columns)
    dd2.loc[0,:] = dd.values.mean(axis=0).compute()
    return dd2
